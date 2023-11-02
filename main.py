from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import openpyxl
import argparse
import cv2


def RGBA2HEX(r: int, g: int, b: int, a=0):
    rh = hex(r)[2:].upper()
    rh = rh if len(rh) != 1 else '0' + rh

    gh = hex(g)[2:].upper()
    gh = gh if len(gh) != 1 else '0' + gh

    bh = hex(b)[2:].upper()
    bh = bh if len(bh) != 1 else '0' + bh

    ah = hex(a)[2:].upper()
    ah = ah if len(ah) != 1 else '0' + ah

    return rh + gh + bh + ah


def image_to_matrix(image_file: str, scale=1.0):
    image = cv2.imread(image_file, cv2.IMREAD_UNCHANGED)
    image = cv2.resize(image, (0, 0), fx=scale, fy=scale)
    image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)

    matrix = []
    for pixel_row in image:
        matrix.append([RGBA2HEX(*pixel_value) for pixel_value in pixel_row])

    return matrix


def image_matrix_to_excel(image_file: str, save_name: str, scale: float):
    image_matrix = image_to_matrix(image_file, scale)
    wb = openpyxl.Workbook()
    ws = wb.active

    # Resize rows height
    for i in range(len(image_matrix)):
        ws.row_dimensions[i +1].width = 1

    # Resize columns width
    for i in range(len(image_matrix[0])):
        ws.column_dimensions[get_column_letter(i +1)].width = 3

    for row, img_row in enumerate(image_matrix, start=1):
        for col, cell_color in enumerate(img_row, start=1):
            ws.cell(row, col).fill = PatternFill(
                start_color=cell_color,
                end_color=cell_color,
                fill_type='solid'
            )

    wb.save(save_name)


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('-i', '--image', help='Path to image', type=str)
    parser.add_argument('-o', '--output', help='Setting output file', type=str)
    parser.add_argument('-s', '--scale', help='Resize output image in excel', type=float)

    args = parser.parse_args()
    image_matrix_to_excel(args.image, args.output, args.scale)
